from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any


class ExcelExporter:
    """Exportador de datos a archivos Excel."""
    
    # Estilos
    HEADER_FILL = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_path: str = None):
        """Inicializa el exportador."""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"Reporte_Ferreteria_{timestamp}.xlsx"
        self.output_path = output_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
    
    def _style_header(self, cell):
        """Aplica estilos a una celda de encabezado."""
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = self.BORDER
    
    def _style_cell(self, cell, alignment="left"):
        """Aplica estilos a una celda de datos."""
        cell.border = self.BORDER
        if alignment == "center":
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif alignment == "right":
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal=alignment, vertical="center")
    
    def add_inventory_sheet(self, products: List[Dict[str, Any]]):
        """Añade hoja de inventario."""
        ws = self.wb.create_sheet("Inventario")

        # Encabezados (actualizados v1.0.2)
        headers = ["ID", "Nombre", "Categoría", "Código", "Proveedor", "Marca", "Unidad",
                  "Precio Compra", "Precio Venta", "Stock", "Stock Mínimo", "Flexible",
                  "Valor Total", "Capital Invertido"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._style_header(cell)

        # Datos
        for row_num, product in enumerate(products, 2):
            stock = float(product.get("stock", 0))
            purchase_price = float(product.get("purchase_price", 0))
            sale_price = float(product.get("sale_price", 0))
            flexible = "SI" if product.get("flexible_stock", False) else "NO"

            ws.cell(row=row_num, column=1, value=product.get("id", ""))
            ws.cell(row=row_num, column=2, value=product.get("name", ""))
            ws.cell(row=row_num, column=3, value=product.get("category", ""))
            ws.cell(row=row_num, column=4, value=product.get("code", ""))
            ws.cell(row=row_num, column=5, value=product.get("provider", ""))
            ws.cell(row=row_num, column=6, value=product.get("marca", ""))
            ws.cell(row=row_num, column=7, value=product.get("unidad", "UNIDAD"))
            ws.cell(row=row_num, column=8, value=purchase_price)
            ws.cell(row=row_num, column=9, value=sale_price)
            ws.cell(row=row_num, column=10, value=stock)
            ws.cell(row=row_num, column=11, value=product.get("min_stock", 0))
            ws.cell(row=row_num, column=12, value=flexible)
            # Valor Total = stock * precio venta
            ws.cell(row=row_num, column=13, value=stock * sale_price)
            # Capital Invertido = stock * precio compra
            ws.cell(row=row_num, column=14, value=stock * purchase_price)

            # Estilos a datos
            for col in range(1, 15):
                cell = ws.cell(row=row_num, column=col)
                self._style_cell(cell, "center" if col in [1, 10, 11, 12] else "left")

                # Formato de precio
                if col in [8, 9, 13, 14]:
                    cell.number_format = "$#,##0.00"

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 16
    
    def add_sales_sheet(self, sales: List[Dict[str, Any]]):
        """Añade hoja de ventas."""
        ws = self.wb.create_sheet("Ventas")

        # Encabezados (actualizados v1.0.2)
        headers = ["ID", "Producto", "Cantidad", "Unidad", "P. Compra", "P. Venta", "Total", "Capital Recup.", "Ganancia", "Fecha", "Hora"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._style_header(cell)

        # Datos
        for row_num, sale in enumerate(sales, 2):
            quantity = float(sale.get("quantity", 0))
            unit_price = float(sale.get("unit_price", 0))
            purchase_price = float(sale.get("purchase_price", 0))
            total = float(sale.get("total", 0))
            # Capital recuperado = quantity * purchase_price
            capital_recovered = quantity * purchase_price
            # Ganancia = total - capital recuperado
            profit = total - capital_recovered

            ws.cell(row=row_num, column=1, value=sale.get("id", ""))
            ws.cell(row=row_num, column=2, value=sale.get("product_name", ""))
            ws.cell(row=row_num, column=3, value=quantity)
            ws.cell(row=row_num, column=4, value=sale.get("unidad", "UNIDAD"))
            ws.cell(row=row_num, column=5, value=purchase_price)
            ws.cell(row=row_num, column=6, value=unit_price)
            ws.cell(row=row_num, column=7, value=total)
            ws.cell(row=row_num, column=8, value=capital_recovered)
            ws.cell(row=row_num, column=9, value=profit)
            ws.cell(row=row_num, column=10, value=sale.get("date", ""))
            ws.cell(row=row_num, column=11, value=sale.get("time", ""))

            # Estilos
            for col in range(1, 12):
                cell = ws.cell(row=row_num, column=col)
                self._style_cell(cell, "center" if col in [1, 3, 10, 11] else "right")

                # Formato de precio
                if col in [5, 6, 7, 8, 9]:
                    cell.number_format = "$#,##0.00"

        # Ajustar ancho
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 10
    
    def add_summary_sheet(self, products: List[Dict[str, Any]], sales: List[Dict[str, Any]]):
        """Añade hoja de resumen con ganancias (actualizada v1.0.2)."""
        ws = self.wb.create_sheet("Resumen", 0)

        # Título
        title_cell = ws.cell(row=1, column=1, value="RESUMEN DE FERRETERÍA")
        title_cell.font = Font(bold=True, size=14, color="2c3e50")
        ws.merge_cells('A1:D1')

        # Fecha de generación
        ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        # ==================== INVENTARIO ====================
        row = 4
        ws.cell(row=row, column=1, value="RESUMEN DE INVENTARIO").font = Font(bold=True, size=12)
        row += 1

        # Calcular totales del inventario
        total_products = len(products)
        total_stock = sum(float(p.get("stock", 0)) for p in products)
        total_capital = sum(float(p.get("purchase_price", 0)) * float(p.get("stock", 0)) for p in products)
        total_value = sum(float(p.get("sale_price", 0)) * float(p.get("stock", 0)) for p in products)
        potential_profit = total_value - total_capital
        low_stock = len([p for p in products if float(p.get("stock", 0)) < float(p.get("min_stock", 0))])

        stats_inventory = [
            ("Total de Productos", total_products),
            ("Stock Total", total_stock),
            ("Capital Invertido", f"${total_capital:,.2f}"),
            ("Valor del Inventario", f"${total_value:,.2f}"),
            ("Ganancia Potencial", f"${potential_profit:,.2f}"),
            ("Productos en Stock Bajo", low_stock),
        ]

        for label, value in stats_inventory:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # ==================== VENTAS ====================
        row += 1
        ws.cell(row=row, column=1, value="RESUMEN DE VENTAS").font = Font(bold=True, size=12)
        row += 1

        total_sales = len(sales)
        total_revenue = sum(float(s.get("total", 0)) for s in sales)
        total_capital_recovered = sum(float(s.get("quantity", 0)) * float(s.get("purchase_price", 0)) for s in sales)
        total_profit = total_revenue - total_capital_recovered

        stats_sales = [
            ("Total de Transacciones", total_sales),
            ("Ingresos Totales", f"${total_revenue:,.2f}"),
            ("Capital Recuperado", f"${total_capital_recovered:,.2f}"),
            ("GANANCIA TOTAL", f"${total_profit:,.2f}"),
        ]

        for label, value in stats_sales:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def save(self) -> bool:
        """Guarda el archivo Excel."""
        try:
            self.wb.save(self.output_path)
            return True
        except Exception as e:
            print(f"Error al guardar Excel: {e}")
            return False
    
    def export(self, products: List[Dict[str, Any]], sales: List[Dict[str, Any]]) -> str:
        """Exporta todos los datos a Excel."""
        try:
            self.add_summary_sheet(products, sales)
            self.add_inventory_sheet(products)
            self.add_sales_sheet(sales)
            
            if self.save():
                return self.output_path
            else:
                return None
        except Exception as e:
            print(f"Error durante exportación: {e}")
            return None
